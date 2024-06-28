from openpyxl import load_workbook
from openpyxl.styles.fonts import Font
import datetime
class Format_tabel:
    def __init__(self):
        super().__init__()


        self.work_days_input()
        self.xlsx_data_sora()
        self.xlsx_data_umi()
        self.xlsx_data_hare()
        self.xlsx_data_sonota()
        # self.main_tabel()


    def work_days_input(self):
        self.days_dict = {
            1: "B",
            2: "C",
            3: "D",
            4: "E",
            5: "F",
            6: "G",
            7: "H",
            8: "I",
            9: "J",
            10: "K",
            11: "L",
            12: "M",
            13: "N",
            14: "O",
            15: "P",
            16: "Q",
            17: "R",
            18: "S",
            19: "T",
            20: "U",
            21: "V",
            22: "W",
            23: "X",
            24: "Y",
            25: "Z",
            26: "AA",
            27: "AB",
            28: "AC",
            29: "AD",
            30: "AE",
            31: "AF",
        }

        # self.input_name = int(input(f"{wb.active}:"))
        self.input_name = 2
        print(self.input_name)

    def xlsx_data_sora(self):

        members = [wsl[f"A{value}"].value for value in range(4, 12)]
        work_day = [wsl[f"{self.days_dict[self.input_name]}{value}"].value for value in range(4, 12)]

        now_data = []
        day_data = []

        for name, day in zip(members, work_day):
            if name and day != "休":
                now_data.append(name)
                day_data.append(day)

        print(now_data)
        print(day_data)
        name_no_data = [y for x in zip(now_data, day_data) for y in x]
        # self.result_sora = " ".join(now_data)
        # self.result_sora = " ".join(name_no_data)
        self.result_sora = now_data
        print(self.result_sora)

    def xlsx_data_umi(self):

        members_2 = [wsl[f"A{value}"].value for value in range(16, 32)]
        work_day_2 = [wsl[f"{self.days_dict[self.input_name]}{value}"].value for value in range(16, 32)]

        now_data = []
        day_data = []

        for name, day in zip(members_2, work_day_2):
            if name and day != "休":
                now_data.append(name)
                day_data.append(day)

        name_no_data = [y for x in zip(now_data, day_data) for y in x]

        # self.result_umi = " ".join(filter(None,name_no_data))
        self.result_umi = now_data
        print(self.result_umi)


    def xlsx_data_hare(self):

        members_3 = [wsl[f"A{value}"].value for value in range(36, 57)]
        work_day_3 = [wsl[f"{self.days_dict[self.input_name]}{value}"].value for value in range(36, 57)]

        now_data = []
        day_data = []

        for name, day in zip(members_3, work_day_3):
            if name and day != "休":
                now_data.append(name)
                day_data.append(day)

        name_no_data = [y for x in zip(now_data, day_data) for y in x]


        # self.result_hare = " ".join(filter(None, name_no_data))
        self.result_hare = now_data
        print(self.result_hare)

    def xlsx_data_sonota(self):

        members_4 = [wsl[f"A{value}"].value for value in range(59, 82)]
        work_day_4 = [wsl[f"{self.days_dict[self.input_name]}{value}"].value for value in range(59, 82)]

        now_data = []
        day_data = []

        for name, day in zip(members_4, work_day_4):
            if name and day != "休":
                now_data.append(name)
                day_data.append(day)

        name_no_data = [y for x in zip(now_data, day_data) for y in x]

        # self.result_sonota = " ".join(filter(None, name_no_data))
        self.result_sonota = now_data
        print(self.result_sonota)

    # def main_tabel(self):
    #     now_day = datetime.datetime.now().date()
    #     print(now_day)
    #
    #     wb = load_workbook('position_table.xlsx')
    #     sheet_names = wb.sheetnames
    #     wsl = wb[sheet_names[0]]
    #     # x = wsl["A4"].value
    #     # ws2 = wb.create_sheet('new_sheet')
    #
    #     wsl['A2'].value = f"2024- 3 - {str(self.input_name)}"
    #     wsl['A2'].font = Font(size=20, bold=True)
    #     wsl['A15'].value = f"空：{self.result_sora}"
    #     wsl['A16'].value = f"海：{self.result_umi}"
    #     wsl['A17'].value = f"晴：{self.result_hare}"
    #     wsl['A18'].value = f"共用・大浴場: {self.result_sonota}"
    #     wsl['D1'].value = f"({str(now_day)})"
    #     wsl['D1'].font = Font(size=20, italic=True)
    #
    #     wb.save('position_table.xlsx')

# if __name__ == '__main__':
#     wb = load_workbook(filename=R"C:\Users\s1non\OneDrive\デスクトップ\月別シフト表コピー.xlsx")
#     wsl = wb[wb.sheetnames[8]]
#
#     app = Format_tabel
#     app()


wb = load_workbook(filename=R"C:\Users\s1non\OneDrive\デスクトップ\月別シフト表コピー.xlsx")
wsl = wb[wb.sheetnames[6]]

app = Format_tabel
app()
